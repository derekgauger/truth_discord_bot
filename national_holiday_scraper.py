"""
This script is used for the 'Truth Discord Bot'. It gathers all the fun national days from the website
'https://nationaldaycalendar.com/' and puts them in a spreadsheet.

In the Truth bot Javascript code, it will read the xlsx file and print all current national days on Discord
for everyone to see.

Pip libraries needed: requests, openpyxl
"""


import requests
import re
import html
from openpyxl import Workbook

months = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]


# Get the HTML and puts it into a text file
def get_webpage(month):
    r = requests.get("https://nationaldaycalendar.com/{}".format(month))

    with open('page.txt', 'w') as f:
        f.write(r.text)


# Gets the holidays from the html file by scraping through the html
def get_holidays(month):
    get_webpage(month)
    
    # Get text file lines
    with open('page.txt') as f:
        lines = f.readlines()

    # Get holidays by checking each line for the holiday entry formats
    matches = []
    for line in lines:
        match = match2 = None
        if "<li>" in line:
            match = re.search(r'<li>.* {} (\d+).*">(.*)</a>.*</li>'.format(month), line)
            match2 = re.search(r'<li>.*{}-(\d+)/">(.*)</a>.*</li>'.format(month.lower()), line)
        if match:
            matches.append(match.groups())
        elif match2:
            matches.append(match2.groups())

    # Create dictionary that contains holidays for each day
    holidays = {}
    for match in matches:
        day = match[0]
        holiday = match[1]
        if day in holidays:
            holidays[day].append(holiday)
        else:
            holidays[day] = [holiday]

    return holidays


# Method for creating a worksheet
def create_worksheet(wb, month):
    ws = wb.create_sheet(month)


# Method for creating the xlsx file from scratch
def create_xlsx_file(wb):
    for month in months:
        create_worksheet(wb, month)

    del wb['Sheet']


# main
def main():
    wb = Workbook()

    create_xlsx_file(wb)

    for month in months:
        ws = wb.get_sheet_by_name(month)
        holidays = get_holidays(month)
        for day in holidays:
            i = 0
            for holiday in holidays[day]:
                holiday = html.unescape(holiday)

                if len(holiday) > 3:
                    ws.cell(row=int(day), column=i + 1).value = holiday
                    i += 1

    wb.save('ze_holidays.xlsx')


main()
